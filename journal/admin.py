import logging

from django.contrib import admin
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _

from .forms import (
    CollegeUserChangeForm,
    CollegeUserCreationForm,
    StudentImportForm,
    StudyGroupAdminForm,
    TeachingAssignmentAdminForm,
)
from .models import (
    AcademicYearArchive,
    AcademicYearArchiveEvent,
    Attendance,
    CollegeInformation,
    CurriculumPlanItem,
    Grade,
    JournalChangeLog,
    Lesson,
    Student,
    StudyGroup,
    Subject,
    TeachingAssignment,
    UserProfile,
    user_full_name,
)
from .reports import build_academic_year_archive, build_report_package
from .permissions import (
    can_download_reports,
    can_manage_academic_structure,
    can_view_all_journals,
)


logger = logging.getLogger(__name__)


admin.site.site_header = "Администрирование электронного журнала"
admin.site.site_title = "Электронный журнал"
admin.site.index_title = "Управление данными"


class TeacherFullNameFilter(admin.SimpleListFilter):
    title = "Преподаватель"
    parameter_name = "teacher_fio"
    relation_path = "teacher_id"

    def lookups(self, request, model_admin):
        user_model = get_user_model()
        teacher_ids = (
            model_admin.get_queryset(request)
            .values_list(self.relation_path, flat=True)
            .distinct()
        )
        users = user_model.objects.filter(pk__in=teacher_ids).select_related(
            "journal_profile"
        )
        return [
            (str(user.pk), user_full_name(user))
            for user in sorted(users, key=lambda item: user_full_name(item))
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(**{self.relation_path: self.value()})
        return queryset


class AssignmentTeacherFullNameFilter(TeacherFullNameFilter):
    relation_path = "assignment__teacher_id"


class AssignmentSubjectFilter(admin.SimpleListFilter):
    title = "Дисциплина"
    parameter_name = "subject"

    def lookups(self, request, model_admin):
        subject_ids = model_admin.get_queryset(request).values_list(
            "assignment__subject_id",
            flat=True,
        )
        subjects = Subject.objects.filter(
            pk__in=subject_ids,
        ).order_by(
            "name"
        )
        return [(str(subject.pk), subject.name) for subject in subjects]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(assignment__subject_id=self.value())
        return queryset


class CollegeUserAdmin(UserAdmin):
    form = CollegeUserChangeForm
    add_form = CollegeUserCreationForm
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        (
            _("Personal info"),
            {
                "fields": (
                    "first_name",
                    "last_name",
                    "middle_name",
                    "email",
                )
            },
        ),
        (
            _("Permissions"),
            {
                "fields": (
                    "is_active",
                    "is_staff",
                    "is_superuser",
                    "groups",
                    "user_permissions",
                ),
            },
        ),
        (_("Important dates"), {"fields": ("last_login", "date_joined")}),
    )
    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": (
                    "username",
                    "password1",
                    "password2",
                    "first_name",
                    "last_name",
                    "middle_name",
                    "email",
                ),
            },
        ),
    )
    list_display = (
        "username",
        "full_name_display",
        "email",
        "is_staff",
        "is_active",
    )
    search_fields = (
        "username",
        "first_name",
        "last_name",
        "journal_profile__middle_name",
        "email",
    )

    @admin.display(description="ФИО", ordering="last_name")
    def full_name_display(self, obj):
        return user_full_name(obj)

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        UserProfile.objects.update_or_create(
            user=obj,
            defaults={
                "middle_name": form.cleaned_data.get(
                    "middle_name",
                    "",
                )
            },
        )


admin.site.unregister(get_user_model())
admin.site.register(get_user_model(), CollegeUserAdmin)


class ArchivedYearProtectionMixin:
    def _academic_year_for_object(self, obj):
        if isinstance(obj, TeachingAssignment):
            return obj.academic_year
        if isinstance(obj, (Lesson, CurriculumPlanItem)):
            return obj.assignment.academic_year
        if isinstance(obj, (Grade, Attendance)):
            return obj.lesson.assignment.academic_year
        return None

    def has_delete_permission(self, request, obj=None):
        if obj is not None:
            academic_year = self._academic_year_for_object(obj)
            if (
                academic_year
                and AcademicYearArchive.is_year_locked(academic_year)
            ):
                return False
        return super().has_delete_permission(request, obj)

    def delete_model(self, request, obj):
        academic_year = self._academic_year_for_object(obj)
        if (
            academic_year
            and AcademicYearArchive.is_year_locked(academic_year)
        ):
            self.message_user(
                request,
                "Объекты архивного учебного года удалять нельзя.",
                level=messages.ERROR,
            )
            return
        super().delete_model(request, obj)

    def delete_queryset(self, request, queryset):
        if any(
            AcademicYearArchive.is_year_locked(
                self._academic_year_for_object(obj)
            )
            for obj in queryset
        ):
            self.message_user(
                request,
                "Выбор содержит данные архивного учебного года.",
                level=messages.ERROR,
            )
            return
        super().delete_queryset(request, queryset)


@admin.register(AcademicYearArchive)
class AcademicYearArchiveAdmin(admin.ModelAdmin):
    list_display = (
        "academic_year",
        "status_display",
        "order_number",
        "order_date",
        "archived_at",
        "archived_by_name",
        "checksum_short",
    )
    search_fields = ("academic_year", "order_number", "checksum")
    list_filter = ("is_locked",)
    actions = (
        "close_academic_year",
        "download_academic_year_archive",
        "reopen_academic_year",
    )
    fields = (
        "academic_year",
        "order_number",
        "order_date",
        "comment",
        "is_locked",
        "archived_at",
        "archived_by_name",
        "archive_filename",
        "checksum",
        "reopened_at",
        "reopened_by_name",
        "reopen_reason",
    )
    readonly_fields = (
        "is_locked",
        "archived_at",
        "archived_by_name",
        "archive_filename",
        "checksum",
        "reopened_at",
        "reopened_by_name",
    )

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        if obj and obj.is_locked:
            readonly.extend(
                [
                    "academic_year",
                    "order_number",
                    "order_date",
                    "comment",
                ]
            )
        return tuple(dict.fromkeys(readonly))

    def has_add_permission(self, request):
        return can_manage_academic_structure(request.user)

    def has_change_permission(self, request, obj=None):
        return can_manage_academic_structure(request.user)

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return can_view_all_journals(request.user)

    def has_module_permission(self, request):
        return can_view_all_journals(request.user)

    @admin.display(description="Статус", ordering="is_locked")
    def status_display(self, obj):
        return "Архивный" if obj.is_locked else "Открыт"

    @admin.display(description="SHA-256")
    def checksum_short(self, obj):
        return f"{obj.checksum[:16]}…" if obj.checksum else "—"

    @admin.display(description="Архивировал")
    def archived_by_name(self, obj):
        return user_full_name(obj.archived_by) or "—"

    @admin.display(description="Повторно открыл")
    def reopened_by_name(self, obj):
        return user_full_name(obj.reopened_by) or "—"

    @admin.action(description="Закрыть год и сформировать архив")
    def close_academic_year(self, request, queryset):
        if queryset.count() != 1:
            self.message_user(
                request,
                "Выберите ровно один учебный год.",
                level=messages.ERROR,
            )
            return None
        archive_record = queryset.get()
        if archive_record.is_locked:
            self.message_user(
                request,
                "Этот учебный год уже закрыт.",
                level=messages.WARNING,
            )
            return None
        if not CollegeInformation.get_current():
            self.message_user(
                request,
                "Сначала заполните сведения об образовательной организации.",
                level=messages.ERROR,
            )
            return None
        if not archive_record.order_number or not archive_record.order_date:
            self.message_user(
                request,
                "Укажите номер и дату приказа об архивировании.",
                level=messages.ERROR,
            )
            return None
        try:
            content, filename, checksum = build_academic_year_archive(
                archive_record.academic_year,
                request.user,
            )
        except ValueError as error:
            self.message_user(request, str(error), level=messages.ERROR)
            return None

        archive_record.archive_data = content
        archive_record.archive_filename = filename
        archive_record.checksum = checksum
        archive_record.is_locked = True
        archive_record.archived_at = timezone.now()
        archive_record.archived_by = request.user
        archive_record.reopened_at = None
        archive_record.reopened_by = None
        archive_record.reopen_reason = ""
        archive_record.save()
        AcademicYearArchiveEvent.objects.create(
            archive=archive_record,
            user=request.user,
            action=AcademicYearArchiveEvent.Action.CLOSE,
            reason=(
                f"Приказ № {archive_record.order_number} "
                f"от {archive_record.order_date:%d.%m.%Y}. "
                f"{archive_record.comment}"
            ).strip(),
            checksum=checksum,
        )
        logger.info(
            "academic_year_closed archive_id=%s academic_year=%s user_id=%s",
            archive_record.pk,
            archive_record.academic_year,
            request.user.pk,
        )
        self.message_user(
            request,
            (
                f"Учебный год {archive_record.academic_year} закрыт. "
                f"SHA-256: {checksum}"
            ),
            level=messages.SUCCESS,
        )

    @admin.action(
        description="Скачать архив учебного года",
        permissions=["view"],
    )
    def download_academic_year_archive(self, request, queryset):
        if queryset.count() != 1:
            self.message_user(
                request,
                "Выберите ровно один учебный год.",
                level=messages.ERROR,
            )
            return None
        archive_record = queryset.get()
        if not archive_record.archive_data:
            self.message_user(
                request,
                "Архив для этого учебного года ещё не сформирован.",
                level=messages.ERROR,
            )
            return None
        response = HttpResponse(
            bytes(archive_record.archive_data),
            content_type="application/zip",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{archive_record.archive_filename}"'
        )
        return response

    @admin.action(description="Повторно открыть учебный год")
    def reopen_academic_year(self, request, queryset):
        if queryset.count() != 1:
            self.message_user(
                request,
                "Выберите ровно один учебный год.",
                level=messages.ERROR,
            )
            return None
        archive_record = queryset.get()
        if not archive_record.is_locked:
            self.message_user(
                request,
                "Этот учебный год уже открыт.",
                level=messages.WARNING,
            )
            return None
        if not archive_record.reopen_reason.strip():
            self.message_user(
                request,
                (
                    "Сначала откройте запись учебного года, укажите причину "
                    "повторного открытия и сохраните её."
                ),
                level=messages.ERROR,
            )
            return None
        archive_record.is_locked = False
        archive_record.reopened_at = timezone.now()
        archive_record.reopened_by = request.user
        archive_record.save(
            update_fields=[
                "is_locked",
                "reopened_at",
                "reopened_by",
                "reopen_reason",
            ]
        )
        AcademicYearArchiveEvent.objects.create(
            archive=archive_record,
            user=request.user,
            action=AcademicYearArchiveEvent.Action.REOPEN,
            reason=archive_record.reopen_reason,
            checksum=archive_record.checksum,
        )
        logger.warning(
            "academic_year_reopened archive_id=%s academic_year=%s user_id=%s",
            archive_record.pk,
            archive_record.academic_year,
            request.user.pk,
        )
        self.message_user(
            request,
            f"Учебный год {archive_record.academic_year} повторно открыт.",
            level=messages.SUCCESS,
        )


@admin.register(AcademicYearArchiveEvent)
class AcademicYearArchiveEventAdmin(admin.ModelAdmin):
    list_display = (
        "created_at",
        "archive",
        "action",
        "user_name",
        "reason",
        "checksum",
    )
    list_filter = ("action", "archive__academic_year", "created_at")
    search_fields = (
        "archive__academic_year",
        "user__username",
        "reason",
        "checksum",
    )
    readonly_fields = (
        "archive",
        "created_at",
        "user_name",
        "action",
        "reason",
        "checksum",
    )

    def has_add_permission(self, request):
        return False

    @admin.display(description="Пользователь")
    def user_name(self, obj):
        return user_full_name(obj.user)

    def has_change_permission(self, request, obj=None):
        return can_view_all_journals(request.user) and request.method in {
            "GET",
            "HEAD",
            "OPTIONS",
        }

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return can_view_all_journals(request.user)

    def has_module_permission(self, request):
        return can_view_all_journals(request.user)


@admin.register(CollegeInformation)
class CollegeInformationAdmin(admin.ModelAdmin):
    fields = ("name", "address", "inn", "email")

    def has_add_permission(self, request):
        return (
            can_manage_academic_structure(request.user)
            and not CollegeInformation.objects.exists()
        )

    def has_change_permission(self, request, obj=None):
        return can_manage_academic_structure(request.user)

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return can_view_all_journals(request.user)

    def has_module_permission(self, request):
        return can_view_all_journals(request.user)


class StudentInline(admin.TabularInline):
    model = Student
    fields = (
        "last_name",
        "first_name",
        "middle_name",
        "student_card_number",
        "is_active",
    )
    readonly_fields = fields
    extra = 0
    max_num = 0
    can_delete = False
    show_change_link = False
    verbose_name = "студент группы"
    verbose_name_plural = "Студенты группы"

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


class ReadOnlyAssignmentInline(admin.TabularInline):
    model = TeachingAssignment
    fields = (
        "subject",
        "group",
        "teacher_full_name",
        "academic_year",
        "semester",
    )
    readonly_fields = fields
    extra = 0
    max_num = 0
    can_delete = False
    show_change_link = False
    verbose_name = "назначение дисциплины группе"

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    @admin.display(description="Преподаватель")
    def teacher_full_name(self, obj):
        return user_full_name(obj.teacher)


class GroupAssignmentInline(ReadOnlyAssignmentInline):
    fields = (
        "subject",
        "teacher_full_name",
        "academic_year",
        "semester",
    )
    readonly_fields = fields
    verbose_name_plural = "Назначенные дисциплины"


class SubjectAssignmentInline(ReadOnlyAssignmentInline):
    fields = (
        "group",
        "teacher_full_name",
        "academic_year",
        "semester",
    )
    readonly_fields = fields
    verbose_name_plural = "Группы и преподаватели дисциплины"


@admin.register(StudyGroup)
class StudyGroupAdmin(admin.ModelAdmin):
    form = StudyGroupAdminForm
    list_display = ("name", "admission_year", "curator_name")
    search_fields = ("name",)
    list_filter = ("admission_year",)
    inlines = (StudentInline, GroupAssignmentInline)

    @admin.display(description="Куратор", ordering="curator__last_name")
    def curator_name(self, obj):
        return user_full_name(obj.curator) or "—"


@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    change_list_template = "admin/journal/student/change_list.html"
    list_display = (
        "last_name",
        "first_name",
        "middle_name",
        "group",
        "student_card_number",
        "is_active",
    )
    list_filter = ("group", "is_active")
    search_fields = (
        "last_name",
        "first_name",
        "middle_name",
        "student_card_number",
    )

    def get_urls(self):
        custom_urls = [
            path(
                "import-xlsx/",
                self.admin_site.admin_view(self.import_xlsx),
                name="journal_student_import",
            ),
        ]
        return custom_urls + super().get_urls()

    def import_xlsx(self, request):
        if not self.has_add_permission(request):
            self.message_user(
                request,
                "У вас нет прав на импорт студентов.",
                level=messages.ERROR,
            )
            return redirect("admin:journal_student_changelist")

        form = StudentImportForm(
            request.POST or None,
            request.FILES or None,
        )
        if request.method == "POST" and form.is_valid():
            try:
                rows = self._read_student_workbook(
                    form.cleaned_data["student_file"]
                )
                created_count, updated_count = self._save_imported_students(
                    form.cleaned_data["group"],
                    rows,
                )
            except ValueError as error:
                form.add_error("student_file", str(error))
            else:
                logger.info(
                    (
                        "students_imported group_id=%s created_count=%s "
                        "updated_count=%s user_id=%s"
                    ),
                    form.cleaned_data["group"].pk,
                    created_count,
                    updated_count,
                    request.user.pk,
                )
                self.message_user(
                    request,
                    (
                        f"Импорт завершён: создано {created_count}, "
                        f"обновлено {updated_count}."
                    ),
                    level=messages.SUCCESS,
                )
                return redirect("admin:journal_student_changelist")

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Импорт студентов из XLSX",
            "form": form,
            "changelist_url": reverse("admin:journal_student_changelist"),
        }
        return TemplateResponse(
            request,
            "admin/journal/student/import_xlsx.html",
            context,
        )

    @staticmethod
    def _normalize_header(value):
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _text_value(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @classmethod
    def _read_student_workbook(cls, uploaded_file):
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(
                uploaded_file,
                data_only=True,
                read_only=True,
            )
        except Exception as error:
            raise ValueError(
                "Не удалось открыть файл. Проверьте, что это корректный XLSX."
            ) from error

        sheet = workbook.active
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            raise ValueError("Файл не содержит данных.")

        aliases = {
            "last_name": {"фамилия"},
            "first_name": {"имя"},
            "middle_name": {"отчество"},
            "student_card_number": {
                "номер студенческого билета",
                "номер студ. билета",
                "студенческий билет",
            },
            "is_active": {"обучается", "активен"},
        }
        headers = [cls._normalize_header(value) for value in all_rows[0]]
        columns = {}
        for field, names in aliases.items():
            for index, header in enumerate(headers):
                if header in names:
                    columns[field] = index
                    break

        required = {"last_name", "first_name", "student_card_number"}
        if not required.issubset(columns):
            raise ValueError(
                "Не найдены обязательные колонки: Фамилия, Имя, "
                "Номер студенческого билета."
            )

        imported_rows = []
        errors = []
        seen_cards = set()
        for excel_row, values in enumerate(all_rows[1:], start=2):
            if not any(value not in (None, "") for value in values):
                continue

            def cell(field):
                index = columns.get(field)
                return (
                    cls._text_value(values[index])
                    if index is not None and index < len(values)
                    else ""
                )

            last_name = cell("last_name")
            first_name = cell("first_name")
            card_number = cell("student_card_number")
            if not last_name or not first_name or not card_number:
                errors.append(
                    f"строка {excel_row}: заполните фамилию, имя и номер билета"
                )
                continue
            middle_name = cell("middle_name")
            if len(last_name) > 100 or len(first_name) > 100:
                errors.append(
                    f"строка {excel_row}: фамилия и имя должны быть "
                    "не длиннее 100 символов"
                )
                continue
            if len(middle_name) > 100:
                errors.append(
                    f"строка {excel_row}: отчество должно быть "
                    "не длиннее 100 символов"
                )
                continue
            if len(card_number) > 30:
                errors.append(
                    f"строка {excel_row}: номер билета должен быть "
                    "не длиннее 30 символов"
                )
                continue
            if card_number in seen_cards:
                errors.append(
                    f"строка {excel_row}: номер билета {card_number} повторяется"
                )
                continue
            seen_cards.add(card_number)

            active_text = cell("is_active").lower()
            is_active = active_text not in {
                "нет",
                "0",
                "false",
                "не обучается",
                "отчислен",
            }
            imported_rows.append(
                {
                    "last_name": last_name,
                    "first_name": first_name,
                    "middle_name": middle_name,
                    "student_card_number": card_number,
                    "is_active": is_active,
                }
            )

        if errors:
            message = "; ".join(errors[:10])
            if len(errors) > 10:
                message += f"; ещё ошибок: {len(errors) - 10}"
            raise ValueError(message)
        if not imported_rows:
            raise ValueError("В файле не найдено ни одного студента.")
        return imported_rows

    @staticmethod
    def _save_imported_students(group, rows):
        created_count = 0
        updated_count = 0
        with transaction.atomic():
            for row in rows:
                card_number = row["student_card_number"]
                _, created = Student.objects.update_or_create(
                    student_card_number=card_number,
                    defaults={
                        "group": group,
                        "last_name": row["last_name"],
                        "first_name": row["first_name"],
                        "middle_name": row["middle_name"],
                        "is_active": row["is_active"],
                    },
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        return created_count, updated_count


@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    change_form_template = "admin/journal/subject/change_form.html"
    list_display = ("name", "short_name")
    search_fields = ("name", "short_name")
    inlines = (SubjectAssignmentInline,)


@admin.register(TeachingAssignment)
class TeachingAssignmentAdmin(ArchivedYearProtectionMixin, admin.ModelAdmin):
    form = TeachingAssignmentAdminForm
    list_display = (
        "subject",
        "group",
        "teacher_name",
        "academic_year",
        "semester",
        "delete_control",
    )
    list_filter = (
        TeacherFullNameFilter,
        "academic_year",
        "semester",
        "group",
        "subject",
    )
    search_fields = (
        "subject__name",
        "group__name",
        "teacher__first_name",
        "teacher__last_name",
        "teacher__journal_profile__middle_name",
        "teacher__username",
    )
    actions = ("download_report_package",)

    def get_actions(self, request):
        actions = super().get_actions(request)
        if not can_download_reports(request.user):
            actions.pop("download_report_package", None)
        return actions

    @admin.display(description="Преподаватель", ordering="teacher__last_name")
    def teacher_name(self, obj):
        return user_full_name(obj.teacher)

    class Media:
        css = {"all": ("admin/css/college_journal_admin.css",)}
        js = ("admin/js/college_journal_admin.js",)

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        subject_id = request.GET.get("subject")
        if subject_id:
            initial["subject"] = subject_id
        return initial

    @admin.action(
        description="Сформировать пакет отчётности PDF + CSV + XLSX",
        permissions=["view"],
    )
    def download_report_package(self, request, queryset):
        if not can_download_reports(request.user):
            self.message_user(
                request,
                "У вас нет прав на формирование отчётности.",
                level="error",
            )
            return None
        if queryset.count() != 1:
            self.message_user(
                request,
                "Выберите ровно одно назначение дисциплины.",
                level="error",
            )
            return None
        assignment = queryset.select_related(
            "group",
            "subject",
            "teacher",
        ).get()
        content, filename = build_report_package(assignment, request.user)
        logger.info(
            "report_package_generated assignment_id=%s user_id=%s",
            assignment.pk,
            request.user.pk,
        )
        response = HttpResponse(content, content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @admin.display(description="Удалить")
    def delete_control(self, obj):
        if AcademicYearArchive.is_year_locked(obj.academic_year):
            return "Архив"
        delete_url = reverse(
            "admin:journal_teachingassignment_quick_delete",
            args=[obj.pk],
        )
        return format_html(
            '<button type="button" class="assignment-delete-control" '
            'data-delete-url="{}" '
            'aria-label="Удалить назначение" '
            'title="Удалить назначение">×</button>',
            delete_url,
        )

    def get_urls(self):
        custom_urls = [
            path(
                "<int:object_id>/quick-delete/",
                self.admin_site.admin_view(self.quick_delete),
                name="journal_teachingassignment_quick_delete",
            )
        ]
        return custom_urls + super().get_urls()

    def quick_delete(self, request, object_id):
        if request.method != "POST":
            return JsonResponse({"detail": "Метод не поддерживается."}, status=405)

        assignment = self.get_object(request, object_id)
        if assignment is None:
            return JsonResponse({"detail": "Назначение не найдено."}, status=404)
        if not self.has_delete_permission(request, assignment):
            return JsonResponse({"detail": "Недостаточно прав."}, status=403)
        if AcademicYearArchive.is_year_locked(assignment.academic_year):
            return JsonResponse(
                {"detail": "Архивный учебный год изменять нельзя."},
                status=403,
            )

        assignment.delete()
        return JsonResponse({"deleted": True})


@admin.register(Lesson)
class LessonAdmin(ArchivedYearProtectionMixin, admin.ModelAdmin):
    list_display = (
        "date",
        "subject_name",
        "topic",
        "lesson_type",
        "teacher_name",
        "group_name",
        "academic_year",
        "semester_name",
    )
    list_filter = (
        "date",
        "assignment__subject",
        AssignmentTeacherFullNameFilter,
        "assignment__group",
        "assignment__academic_year",
        "assignment__semester",
        "lesson_type",
        "grading_scheme",
    )
    search_fields = (
        "topic",
        "assignment__subject__name",
        "assignment__group__name",
        "assignment__teacher__first_name",
        "assignment__teacher__last_name",
        "assignment__teacher__journal_profile__middle_name",
        "assignment__teacher__username",
    )
    list_select_related = (
        "assignment__subject",
        "assignment__teacher",
        "assignment__group",
    )
    date_hierarchy = "date"

    @admin.display(
        description="Дисциплина",
        ordering="assignment__subject__name",
    )
    def subject_name(self, obj):
        return obj.assignment.subject

    @admin.display(
        description="Преподаватель",
        ordering="assignment__teacher__last_name",
    )
    def teacher_name(self, obj):
        return user_full_name(obj.assignment.teacher)

    @admin.display(
        description="Группа",
        ordering="assignment__group__name",
    )
    def group_name(self, obj):
        return obj.assignment.group

    @admin.display(
        description="Учебный год",
        ordering="assignment__academic_year",
    )
    def academic_year(self, obj):
        return obj.assignment.academic_year

    @admin.display(
        description="Семестр",
        ordering="assignment__semester",
    )
    def semester_name(self, obj):
        return obj.assignment.get_semester_display()


@admin.register(Grade)
class GradeAdmin(ArchivedYearProtectionMixin, admin.ModelAdmin):
    list_display = ("student", "lesson", "value", "pass_result")
    list_filter = ("value", "lesson__assignment__group", "lesson__assignment__subject")
    search_fields = ("student__last_name", "student__first_name")


@admin.register(CurriculumPlanItem)
class CurriculumPlanItemAdmin(
    ArchivedYearProtectionMixin,
    admin.ModelAdmin,
):
    list_display = (
        "sequence",
        "topic",
        "lesson_type",
        "hours",
        "assignment",
    )
    list_filter = (
        AssignmentSubjectFilter,
        AssignmentTeacherFullNameFilter,
        "assignment__academic_year",
        "assignment__semester",
        "assignment__group",
        "lesson_type",
    )
    search_fields = (
        "topic",
        "assignment__subject__name",
        "assignment__group__name",
        "assignment__teacher__first_name",
        "assignment__teacher__last_name",
        "assignment__teacher__journal_profile__middle_name",
    )
    list_select_related = (
        "assignment__subject",
        "assignment__group",
        "assignment__teacher",
    )


@admin.register(Attendance)
class AttendanceAdmin(ArchivedYearProtectionMixin, admin.ModelAdmin):
    list_display = ("student", "lesson", "status")
    list_filter = (
        "status",
        "lesson__assignment__group",
        "lesson__assignment__subject",
    )
    search_fields = ("student__last_name", "student__first_name")


@admin.register(JournalChangeLog)
class JournalChangeLogAdmin(admin.ModelAdmin):
    list_display = (
        "created_at",
        "user_name",
        "assignment",
        "lesson_date",
        "student_name",
        "entity_type",
        "action",
        "old_value",
        "new_value",
    )
    list_filter = (
        "entity_type",
        "action",
        "assignment__group",
        "assignment__subject",
        "created_at",
    )
    search_fields = (
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__journal_profile__middle_name",
        "student_name",
        "lesson_topic",
        "reason",
        "comment",
    )
    readonly_fields = (
        "created_at",
        "user_name",
        "assignment",
        "lesson_date",
        "lesson_topic",
        "student",
        "student_name",
        "entity_type",
        "action",
        "old_value",
        "new_value",
        "reason",
        "comment",
    )

    def has_add_permission(self, request):
        return False

    @admin.display(description="Пользователь")
    def user_name(self, obj):
        return user_full_name(obj.user)

    def has_change_permission(self, request, obj=None):
        return can_view_all_journals(request.user) and request.method in {
            "GET",
            "HEAD",
            "OPTIONS",
        }

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return can_view_all_journals(request.user)

    def has_module_permission(self, request):
        return can_view_all_journals(request.user)
