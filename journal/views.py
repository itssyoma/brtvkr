from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import logging
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Avg, Prefetch, Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .models import (
    AcademicYearArchive,
    Attendance,
    CurriculumPlanItem,
    Grade,
    GradingScheme,
    JournalChangeLog,
    Lesson,
    LessonType,
    Student,
    TeachingAssignment,
    user_full_name,
)
from .permissions import can_view_all_journals, is_system_administrator


logger = logging.getLogger(__name__)


def home(request):
    return render(request, "journal/home.html")


def _available_assignments(user):
    assignments = TeachingAssignment.objects.select_related(
        "teacher",
        "group",
        "subject",
    )
    if is_system_administrator(user):
        return assignments
    return assignments.filter(teacher=user)


def _curated_assignments(user):
    if is_system_administrator(user):
        return TeachingAssignment.objects.none()
    return (
        TeachingAssignment.objects.select_related(
            "teacher",
            "group",
            "subject",
        )
        .filter(group__curator=user)
        .exclude(teacher=user)
    )


def _viewable_assignments(user):
    if can_view_all_journals(user):
        return TeachingAssignment.objects.select_related(
            "teacher",
            "group",
            "subject",
        )
    return TeachingAssignment.objects.select_related(
        "teacher",
        "group",
        "subject",
    ).filter(
        Q(teacher=user) | Q(group__curator=user)
    )


def _get_available_assignment(user, assignment_id):
    return get_object_or_404(_available_assignments(user), pk=assignment_id)


def _get_viewable_assignment(user, assignment_id):
    return get_object_or_404(_viewable_assignments(user), pk=assignment_id)


def _can_edit_assignment(user, assignment):
    return (
        is_system_administrator(user)
        or assignment.teacher_id == user.id
    )


def _assignment_is_archived(assignment):
    return AcademicYearArchive.is_year_locked(assignment.academic_year)


def _reject_archived_change(request, assignment):
    if not _assignment_is_archived(assignment):
        return None
    logger.warning(
        "archived_change_rejected assignment_id=%s user_id=%s",
        assignment.pk,
        request.user.pk,
    )
    messages.error(
        request,
        (
            f"Учебный год {assignment.academic_year} находится в архиве. "
            "Изменение данных запрещено."
        ),
    )
    return redirect(
        "journal:assignment_detail",
        assignment_id=assignment.pk,
    )


def _lesson_type_from_text(value):
    normalized = re.sub(r"\s+", " ", str(value or "").strip().lower())
    mappings = [
        ("дифф", LessonType.DIFFERENTIATED_CREDIT),
        ("семестров", LessonType.SEMESTER_GRADE),
        ("консультац", LessonType.CONSULTATION),
        ("экзамен", LessonType.EXAM),
        ("зач", LessonType.CREDIT),
        ("лаборатор", LessonType.LABORATORY),
        ("практич", LessonType.PRACTICE),
        ("самостоятель", LessonType.SELF_STUDY),
        ("лекц", LessonType.LECTURE),
    ]
    for marker, lesson_type in mappings:
        if marker in normalized:
            return lesson_type
    return LessonType.LECTURE


def _grading_scheme_for_lesson_type(lesson_type):
    if lesson_type == LessonType.CREDIT:
        return GradingScheme.PASS_FAIL
    if lesson_type == LessonType.CONSULTATION:
        return GradingScheme.NONE
    return GradingScheme.NUMERIC


def _grade_display(grade):
    if grade is None:
        return ""
    if grade.pass_result == Grade.PassResult.PASSED:
        return "Зачёт"
    if grade.pass_result == Grade.PassResult.FAILED:
        return "Незачёт"
    return str(grade.value) if grade.value is not None else ""


def _write_change_log(
    *,
    user,
    student,
    lesson,
    entity_type,
    old_value,
    new_value,
    reason,
    comment,
):
    if old_value == new_value:
        return
    action = JournalChangeLog.Action.UPDATE
    if not old_value and new_value:
        action = JournalChangeLog.Action.CREATE
    elif old_value and not new_value:
        action = JournalChangeLog.Action.DELETE
    JournalChangeLog.objects.create(
        user=user,
        assignment=lesson.assignment,
        lesson_date=lesson.date,
        lesson_topic=lesson.topic,
        student=student,
        student_name=str(student),
        entity_type=entity_type,
        action=action,
        old_value=old_value,
        new_value=new_value,
        reason=reason,
        comment=comment,
    )


class ChangeReasonRequired(ValueError):
    pass


def _change_reason_from_request(request):
    reason = request.POST.get("change_reason", "").strip()
    if reason == "__other__":
        return request.POST.get("change_reason_other", "").strip()
    return reason


def _save_grade_result(
    student,
    lesson,
    raw_value,
    *,
    user,
    reason,
    comment="",
    require_reason=None,
):
    value = str(raw_value or "").strip().upper()
    existing = Grade.objects.filter(student=student, lesson=lesson).first()
    old_value = _grade_display(existing)
    new_value = ""

    if lesson.grading_scheme == GradingScheme.NONE:
        new_value = ""
    elif lesson.grading_scheme == GradingScheme.PASS_FAIL:
        pass_result = {
            "З": Grade.PassResult.PASSED,
            "ЗАЧ": Grade.PassResult.PASSED,
            "PASSED": Grade.PassResult.PASSED,
            "НЗ": Grade.PassResult.FAILED,
            "НЕЗАЧ": Grade.PassResult.FAILED,
            "FAILED": Grade.PassResult.FAILED,
        }.get(value)
        if pass_result:
            new_value = (
                "Зачёт"
                if pass_result == Grade.PassResult.PASSED
                else "Незачёт"
            )
        elif value == "":
            new_value = ""
        else:
            raise ValueError("Недопустимый результат зачёта.")
    elif value in {"2", "3", "4", "5"}:
        new_value = value
    elif value == "":
        new_value = ""
    else:
        raise ValueError("Недопустимая оценка.")

    if old_value == new_value:
        return
    if require_reason is None:
        require_reason = existing is not None
    if require_reason and not reason:
        raise ChangeReasonRequired("Укажите основание изменения оценки.")

    if not new_value:
        Grade.objects.filter(student=student, lesson=lesson).delete()
    elif lesson.grading_scheme == GradingScheme.PASS_FAIL:
        pass_result = (
            Grade.PassResult.PASSED
            if new_value == "Зачёт"
            else Grade.PassResult.FAILED
        )
        Grade.objects.update_or_create(
            student=student,
            lesson=lesson,
            defaults={"value": None, "pass_result": pass_result},
        )
    else:
        Grade.objects.update_or_create(
            student=student,
            lesson=lesson,
            defaults={"value": int(new_value), "pass_result": ""},
        )

    _write_change_log(
        user=user,
        student=student,
        lesson=lesson,
        entity_type=JournalChangeLog.EntityType.GRADE,
        old_value=old_value,
        new_value=new_value,
        reason=reason or "Первичное внесение оценки",
        comment=comment,
    )


def _save_attendance(
    student,
    lesson,
    raw_value,
    *,
    user,
    reason,
    comment="",
    require_reason=None,
):
    raw_text = str(raw_value or "").strip()
    if raw_text in Attendance.Status.values:
        status = raw_text
    else:
        status = {
        "": Attendance.Status.PRESENT,
        "Н": Attendance.Status.ABSENT,
        "N": Attendance.Status.ABSENT,
        "У": Attendance.Status.EXCUSED,
        "U": Attendance.Status.EXCUSED,
        "О": Attendance.Status.LATE,
        "O": Attendance.Status.LATE,
        }.get(raw_text.upper())
    if status is None:
        raise ValueError("Недопустимая отметка посещаемости.")
    existing = Attendance.objects.filter(
        student=student,
        lesson=lesson,
    ).first()
    old_value = existing.get_status_display() if existing else ""
    if existing is None and status == Attendance.Status.PRESENT:
        return
    new_value = Attendance.Status(status).label
    if old_value == new_value:
        return
    existing_has_mark = (
        existing is not None
        and existing.status != Attendance.Status.PRESENT
    )
    if require_reason is None:
        require_reason = existing_has_mark
    else:
        require_reason = require_reason and existing_has_mark
    if require_reason and not reason:
        raise ChangeReasonRequired("Укажите основание изменения посещаемости.")
    current, _ = Attendance.objects.update_or_create(
        student=student,
        lesson=lesson,
        defaults={"status": status},
    )
    _write_change_log(
        user=user,
        student=student,
        lesson=lesson,
        entity_type=JournalChangeLog.EntityType.ATTENDANCE,
        old_value=old_value,
        new_value=current.get_status_display(),
        reason=reason or "Первичное внесение посещаемости",
        comment=comment,
    )


def _parse_curriculum_workbook(uploaded_file):
    from openpyxl import load_workbook

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    header_row = None
    columns = {}

    for row_number in range(1, min(worksheet.max_row, 30) + 1):
        values = [
            str(worksheet.cell(row_number, column).value or "").strip().lower()
            for column in range(1, worksheet.max_column + 1)
        ]
        if any("тема" in value for value in values):
            header_row = row_number
            for index, value in enumerate(values, start=1):
                if "№" in value or "п/п" in value:
                    columns["sequence"] = index
                elif "дата" in value:
                    columns["planned_date"] = index
                elif "тема" in value:
                    columns["topic"] = index
                elif "вид" in value:
                    columns["lesson_type"] = index
                elif "час" in value:
                    columns["hours"] = index
                elif "литератур" in value:
                    columns["literature"] = index
            break

    if header_row is None or "topic" not in columns:
        raise ValueError("Не удалось найти столбец с темами занятий.")

    items = []
    total_hours = Decimal("0")
    next_sequence = 1

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        raw_topic = worksheet.cell(row_number, columns["topic"]).value
        raw_type = (
            worksheet.cell(row_number, columns.get("lesson_type", 0)).value
            if columns.get("lesson_type")
            else None
        )
        raw_date = (
            worksheet.cell(row_number, columns.get("planned_date", 0)).value
            if columns.get("planned_date")
            else None
        )
        raw_hours = (
            worksheet.cell(row_number, columns.get("hours", 0)).value
            if columns.get("hours")
            else 1
        )
        raw_literature = (
            worksheet.cell(row_number, columns.get("literature", 0)).value
            if columns.get("literature")
            else ""
        )
        first_cells = [
            worksheet.cell(row_number, column).value
            for column in range(1, min(worksheet.max_column, 4) + 1)
        ]
        row_text = " ".join(str(value or "") for value in first_cells).strip()

        normalized_row_text = re.sub(
            r"\s+",
            " ",
            row_text.lower(),
        ).strip(" :")
        if normalized_row_text == "итого":
            continue

        topic = str(raw_topic or "").strip()
        type_source = raw_type
        if not topic:
            special_type = _lesson_type_from_text(row_text)
            if special_type in {
                LessonType.CONSULTATION,
                LessonType.EXAM,
                LessonType.CREDIT,
                LessonType.DIFFERENTIATED_CREDIT,
                LessonType.SEMESTER_GRADE,
            }:
                topic = LessonType(special_type).label
                type_source = row_text
            else:
                continue

        raw_sequence = (
            worksheet.cell(row_number, columns.get("sequence", 0)).value
            if columns.get("sequence")
            else None
        )
        match = re.search(r"\d+", str(raw_sequence or ""))
        sequence = int(match.group()) if match else next_sequence
        next_sequence = max(next_sequence, sequence + 1)

        planned_date = None
        if isinstance(raw_date, (date, datetime)):
            planned_date = raw_date.date().isoformat() if isinstance(raw_date, datetime) else raw_date.isoformat()
        elif raw_date:
            for date_format in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
                try:
                    planned_date = datetime.strptime(
                        str(raw_date).strip(),
                        date_format,
                    ).date().isoformat()
                    break
                except ValueError:
                    continue

        try:
            hours = Decimal(str(raw_hours or 1).replace(",", "."))
        except InvalidOperation:
            hours = Decimal("1")
        total_hours += hours
        lesson_type = _lesson_type_from_text(type_source)

        items.append(
            {
                "sequence": sequence,
                "planned_date": planned_date,
                "topic": topic[:500],
                "lesson_type": lesson_type,
                "lesson_type_label": LessonType(lesson_type).label,
                "hours": str(hours),
                "literature": str(raw_literature or "").strip()[:500],
            }
        )

    if not items:
        raise ValueError("В файле не найдено ни одной темы КТП.")
    return items, str(total_hours)


@login_required
def dashboard(request):
    assignments = _available_assignments(request.user).order_by(
        "group__name",
        "subject__name",
    )
    curated_assignments = _curated_assignments(request.user).order_by(
        "group__name",
        "subject__name",
    )
    for assignment in curated_assignments:
        assignment.teacher_display_name = user_full_name(assignment.teacher)
    oversight_assignments = TeachingAssignment.objects.none()
    if (
        can_view_all_journals(request.user)
        and not is_system_administrator(request.user)
    ):
        oversight_assignments = TeachingAssignment.objects.select_related(
            "teacher",
            "group",
            "subject",
        ).order_by(
            "group__name",
            "subject__name",
        )
        for assignment in oversight_assignments:
            assignment.teacher_display_name = user_full_name(
                assignment.teacher
            )
    return render(
        request,
        "journal/dashboard.html",
        {
            "assignments": assignments,
            "curated_assignments": curated_assignments,
            "oversight_assignments": oversight_assignments,
            "user_display_name": user_full_name(request.user),
        },
    )


@login_required
def assignment_detail(request, assignment_id):
    assignment = _get_viewable_assignment(request.user, assignment_id)
    can_edit = _can_edit_assignment(request.user, assignment)
    is_archived = _assignment_is_archived(assignment)

    if request.method == "POST":
        if not can_edit:
            raise PermissionDenied
        archived_response = _reject_archived_change(request, assignment)
        if archived_response:
            return archived_response
        action = request.POST.get("action")

        if action == "save_journal":
            reason = _change_reason_from_request(request)
            comment = request.POST.get("change_comment", "").strip()
            lessons_to_update = list(assignment.lessons.all())
            students_to_update = list(
                Student.objects.filter(
                    group=assignment.group,
                    is_active=True,
                )
            )
            try:
                with transaction.atomic():
                    for student in students_to_update:
                        for lesson in lessons_to_update:
                            grade_field = f"grade_{student.pk}_{lesson.pk}"
                            grade_value = request.POST.get(
                                grade_field,
                                "",
                            ).strip()
                            _save_grade_result(
                                student,
                                lesson,
                                grade_value,
                                user=request.user,
                                reason=reason,
                                comment=comment,
                            )

                            attendance_field = (
                                f"attendance_{student.pk}_{lesson.pk}"
                            )
                            attendance_mark = (
                                request.POST.get(attendance_field, "")
                                .strip()
                                .upper()
                            )
                            _save_attendance(
                                student,
                                lesson,
                                attendance_mark,
                                user=request.user,
                                reason=reason,
                                comment=comment,
                            )
            except ChangeReasonRequired as error:
                messages.error(request, str(error))
                return redirect(f"{request.path}?edit=1")
            messages.success(
                request,
                "Оценки и посещаемость в журнале сохранены.",
            )
            return redirect("journal:assignment_detail", assignment_id=assignment.pk)

        curriculum_item_id = request.POST.get("curriculum_item", "").strip()
        curriculum_item = None
        if curriculum_item_id:
            curriculum_item = get_object_or_404(
                CurriculumPlanItem,
                pk=curriculum_item_id,
                assignment=assignment,
            )
        topic = (
            curriculum_item.topic
            if curriculum_item
            else request.POST.get("topic", "").strip()
        )
        lesson_date = request.POST.get("date", "").strip()
        if not topic or not lesson_date:
            messages.error(request, "Укажите дату и тему занятия.")
        else:
            lesson = Lesson.objects.create(
                assignment=assignment,
                date=lesson_date,
                topic=topic,
                curriculum_item=curriculum_item,
                lesson_type=curriculum_item.lesson_type
                if curriculum_item
                else LessonType.LECTURE,
                grading_scheme=_grading_scheme_for_lesson_type(
                    curriculum_item.lesson_type
                    if curriculum_item
                    else LessonType.LECTURE
                ),
            )
            logger.info(
                "lesson_created assignment_id=%s lesson_id=%s user_id=%s",
                assignment.pk,
                lesson.pk,
                request.user.pk,
            )
            messages.success(request, "Занятие добавлено.")
            assignment_url = redirect(
                "journal:assignment_detail",
                assignment_id=assignment.pk,
            )
            if request.POST.get("return_to_edit") == "1":
                assignment_url["Location"] += "?edit=1"
            return assignment_url

    lessons = list(
        assignment.lessons.prefetch_related(
            "grades",
            "attendance_records",
        )
    )
    curriculum_items = list(
        assignment.curriculum_items.prefetch_related("lessons").all()
    )
    for curriculum_item in curriculum_items:
        curriculum_item.is_used = bool(curriculum_item.lessons.all())
    students = list(
        Student.objects.filter(
            group=assignment.group,
            is_active=True,
        )
    )

    grades_by_student_lesson = {
        (grade.student_id, lesson.id): grade
        for lesson in lessons
        for grade in lesson.grades.all()
    }
    attendance_by_student_lesson = {
        (record.student_id, lesson.id): record
        for lesson in lessons
        for record in lesson.attendance_records.all()
    }
    attendance_marks = {
        Attendance.Status.PRESENT: "",
        Attendance.Status.ABSENT: "Н",
        Attendance.Status.EXCUSED: "У",
        Attendance.Status.LATE: "О",
    }

    journal_rows = []
    for student in students:
        cells = []
        grade_values = []
        absence_count = 0

        for lesson in reversed(lessons):
            grade_record = grades_by_student_lesson.get((student.id, lesson.id))
            grade_value = grade_record.value if grade_record else None
            pass_result = grade_record.pass_result if grade_record else ""
            grade_display = grade_value
            grade_input = str(grade_value) if grade_value is not None else ""
            if pass_result == Grade.PassResult.PASSED:
                grade_display = "З"
                grade_input = "З"
            elif pass_result == Grade.PassResult.FAILED:
                grade_display = "НЗ"
                grade_input = "НЗ"
            attendance_record = attendance_by_student_lesson.get(
                (student.id, lesson.id)
            )
            attendance = (
                attendance_record.status
                if attendance_record
                else Attendance.Status.PRESENT
            )
            if grade_value is not None:
                grade_values.append(grade_value)
            if attendance in {
                Attendance.Status.ABSENT,
                Attendance.Status.EXCUSED,
            }:
                absence_count += 1
            cells.append(
                {
                    "lesson": lesson,
                    "grade": grade_display,
                    "grade_input": grade_input,
                    "grade_exists": grade_record is not None,
                    "grading_scheme": lesson.grading_scheme,
                    "attendance": attendance,
                    "attendance_mark": attendance_marks[attendance],
                    "attendance_exists": attendance_record is not None,
                }
            )

        journal_rows.append(
            {
                "student": student,
                "cells": cells,
                "average": round(sum(grade_values) / len(grade_values), 2)
                if grade_values
                else None,
                "absence_count": absence_count,
            }
        )

    return render(
        request,
        "journal/assignment_detail.html",
        {
            "assignment": assignment,
            "lessons": lessons,
            "journal_lessons": list(reversed(lessons)),
            "journal_rows": journal_rows,
            "curriculum_items": curriculum_items,
            "edit_mode": (
                request.GET.get("edit") == "1"
                and can_edit
                and not is_archived
            ),
            "can_edit": can_edit,
            "is_archived": is_archived,
            "today": date.today().isoformat(),
        },
    )


@login_required
def import_curriculum(request, assignment_id):
    assignment = _get_available_assignment(request.user, assignment_id)
    is_archived = _assignment_is_archived(assignment)
    if request.method == "POST":
        archived_response = _reject_archived_change(request, assignment)
        if archived_response:
            return archived_response
    session_key = f"curriculum_import_{assignment.pk}"
    preview = request.session.get(session_key)

    if request.method == "POST" and request.POST.get("action") == "confirm":
        if not preview:
            messages.error(request, "Сначала загрузите и проверьте файл КТП.")
        else:
            with transaction.atomic():
                imported_sequences = set()
                for item in preview["items"]:
                    imported_sequences.add(item["sequence"])
                    CurriculumPlanItem.objects.update_or_create(
                        assignment=assignment,
                        sequence=item["sequence"],
                        defaults={
                            "planned_date": item["planned_date"],
                            "topic": item["topic"],
                            "lesson_type": item["lesson_type"],
                            "hours": Decimal(item["hours"]),
                            "literature": item["literature"],
                        },
                    )
                assignment.curriculum_items.exclude(
                    sequence__in=imported_sequences
                ).filter(lessons__isnull=True).delete()
            del request.session[session_key]
            logger.info(
                "curriculum_imported assignment_id=%s item_count=%s user_id=%s",
                assignment.pk,
                len(preview["items"]),
                request.user.pk,
            )
            messages.success(
                request,
                f"КТП импортирован: {len(preview['items'])} записей.",
            )
            return redirect(
                "journal:assignment_detail",
                assignment_id=assignment.pk,
            )

    if request.method == "POST" and request.FILES.get("curriculum_file"):
        uploaded_file = request.FILES["curriculum_file"]
        if not uploaded_file.name.lower().endswith(".xlsx"):
            messages.error(request, "Поддерживаются только файлы .xlsx.")
        else:
            try:
                items, total_hours = _parse_curriculum_workbook(uploaded_file)
                preview = {
                    "filename": uploaded_file.name,
                    "items": items,
                    "total_hours": total_hours,
                }
                request.session[session_key] = preview
            except (ValueError, OSError) as error:
                messages.error(request, str(error))

    return render(
        request,
        "journal/import_curriculum.html",
        {
            "assignment": assignment,
            "preview": preview,
            "is_archived": is_archived,
        },
    )


@login_required
def curriculum_plan(request, assignment_id):
    assignment = _get_available_assignment(request.user, assignment_id)
    is_archived = _assignment_is_archived(assignment)
    items = list(
        assignment.curriculum_items.prefetch_related("lessons").all()
    )

    if request.method == "POST":
        archived_response = _reject_archived_change(request, assignment)
        if archived_response:
            return archived_response
        action = request.POST.get("action")
        if action == "add":
            last_sequence = (
                assignment.curriculum_items.order_by("-sequence")
                .values_list("sequence", flat=True)
                .first()
                or 0
            )
            try:
                insert_sequence = int(
                    request.POST.get("insert_sequence", last_sequence + 1)
                )
            except (TypeError, ValueError):
                insert_sequence = 0

            if insert_sequence < 1 or insert_sequence > last_sequence + 1:
                messages.error(
                    request,
                    f"Укажите номер от 1 до {last_sequence + 1}.",
                )
            else:
                with transaction.atomic():
                    rows_to_shift = list(
                        assignment.curriculum_items.filter(
                        sequence__gte=insert_sequence
                        ).order_by("-sequence")
                    )
                    for row in rows_to_shift:
                        row.sequence += 1
                        row.save(update_fields=["sequence"])
                    CurriculumPlanItem.objects.create(
                        assignment=assignment,
                        sequence=insert_sequence,
                        topic="Новая тема",
                        lesson_type=LessonType.LECTURE,
                        hours=1,
                    )
                messages.success(
                    request,
                    f"Новая строка добавлена под номером {insert_sequence}.",
                )
            return redirect(
                "journal:curriculum_plan",
                assignment_id=assignment.pk,
            )

        if action == "save":
            submitted_rows = []
            for item in items:
                sequence_raw = request.POST.get(
                    f"sequence_{item.pk}",
                    str(item.sequence),
                ).strip()
                topic = request.POST.get(f"topic_{item.pk}", "").strip()
                lesson_type = request.POST.get(
                    f"lesson_type_{item.pk}",
                    item.lesson_type,
                )
                hours_raw = request.POST.get(
                    f"hours_{item.pk}",
                    str(item.hours),
                ).replace(",", ".")
                planned_date = request.POST.get(
                    f"planned_date_{item.pk}",
                    "",
                ) or None
                literature = request.POST.get(
                    f"literature_{item.pk}",
                    "",
                ).strip()

                try:
                    sequence = int(sequence_raw)
                    hours = Decimal(hours_raw)
                except (ValueError, InvalidOperation):
                    messages.error(
                        request,
                        "Номер и количество часов должны быть числами.",
                    )
                    break
                if sequence < 1 or hours <= 0 or not topic:
                    messages.error(
                        request,
                        "Укажите положительный номер, часы и название темы.",
                    )
                    break
                if lesson_type not in LessonType.values:
                    messages.error(request, "Выбран неизвестный вид занятия.")
                    break
                submitted_rows.append(
                    {
                        "item": item,
                        "sequence": sequence,
                        "topic": topic,
                        "lesson_type": lesson_type,
                        "hours": hours,
                        "planned_date": planned_date,
                        "literature": literature,
                    }
                )
            else:
                # A changed number means "move this row to that position".
                # Keep submitted table order as a stable tie-breaker.
                submitted_rows.sort(
                    key=lambda row: (
                        row["sequence"],
                        row["sequence"] == row["item"].sequence,
                        row["item"].sequence,
                        row["item"].pk,
                    )
                )
                with transaction.atomic():
                    offset = 1_000_000
                    for index, row in enumerate(submitted_rows, start=1):
                        row["item"].sequence = offset + index
                        row["item"].save(update_fields=["sequence"])
                    for index, row in enumerate(submitted_rows, start=1):
                        item = row["item"]
                        item.sequence = index
                        item.topic = row["topic"][:500]
                        item.lesson_type = row["lesson_type"]
                        item.hours = row["hours"]
                        item.planned_date = row["planned_date"]
                        item.literature = row["literature"][:500]
                        item.save()
                messages.success(
                    request,
                    "Изменения КТП сохранены, строки перенумерованы.",
                )
                return redirect(
                    "journal:curriculum_plan",
                    assignment_id=assignment.pk,
                )

    for item in items:
        item.is_used = bool(item.lessons.all())
        item.actual_dates = sorted(
            {lesson.date for lesson in item.lessons.all()}
        )
    next_sequence = (
        assignment.curriculum_items.order_by("-sequence")
        .values_list("sequence", flat=True)
        .first()
        or 0
    ) + 1

    return render(
        request,
        "journal/curriculum_plan.html",
        {
            "assignment": assignment,
            "items": items,
            "next_sequence": next_sequence,
            "lesson_type_choices": LessonType.choices,
            "is_archived": is_archived,
        },
    )


@login_required
@require_POST
def delete_curriculum_item(request, assignment_id, item_id):
    assignment = _get_available_assignment(request.user, assignment_id)
    archived_response = _reject_archived_change(request, assignment)
    if archived_response:
        return archived_response
    item = get_object_or_404(
        CurriculumPlanItem,
        pk=item_id,
        assignment=assignment,
    )
    if item.lessons.exists():
        messages.error(
            request,
            "Нельзя удалить тему, по которой уже создано занятие.",
        )
    else:
        deleted_sequence = item.sequence
        with transaction.atomic():
            item.delete()
            rows_to_shift = list(
                assignment.curriculum_items.filter(
                    sequence__gt=deleted_sequence
                ).order_by("sequence")
            )
            for row in rows_to_shift:
                row.sequence -= 1
                row.save(update_fields=["sequence"])
        messages.success(request, "Строка КТП удалена.")
    return redirect(
        "journal:curriculum_plan",
        assignment_id=assignment.pk,
    )


@login_required
@require_POST
def autosave_journal_cell(request, assignment_id):
    assignment = _get_available_assignment(request.user, assignment_id)
    if _assignment_is_archived(assignment):
        return JsonResponse(
            {
                "error": (
                    f"Учебный год {assignment.academic_year} находится "
                    "в архиве."
                )
            },
            status=403,
        )
    student = get_object_or_404(
        Student,
        pk=request.POST.get("student_id"),
        group=assignment.group,
        is_active=True,
    )
    lesson = get_object_or_404(
        Lesson,
        pk=request.POST.get("lesson_id"),
        assignment=assignment,
    )
    field = request.POST.get("field")
    value = request.POST.get("value", "").strip().upper()
    reason = request.POST.get("reason", "").strip()
    comment = request.POST.get("comment", "").strip()
    existed_before_edit = (
        request.POST.get("original_exists", "").lower() == "true"
    )

    if field == "grade":
        try:
            _save_grade_result(
                student,
                lesson,
                value,
                user=request.user,
                reason=reason,
                comment=comment,
                require_reason=existed_before_edit,
            )
        except ChangeReasonRequired as error:
            return JsonResponse(
                {"error": str(error), "reason_required": True},
                status=400,
            )
        except ValueError as error:
            return JsonResponse({"error": str(error)}, status=400)
    elif field == "attendance":
        try:
            _save_attendance(
                student,
                lesson,
                value,
                user=request.user,
                reason=reason,
                comment=comment,
                require_reason=existed_before_edit,
            )
        except ChangeReasonRequired as error:
            return JsonResponse(
                {"error": str(error), "reason_required": True},
                status=400,
            )
        except ValueError as error:
            return JsonResponse(
                {"error": str(error)},
                status=400,
            )
    else:
        return JsonResponse({"error": "Неизвестное поле."}, status=400)

    student_grades = Grade.objects.filter(
        student=student,
        lesson__assignment=assignment,
        value__isnull=False,
    )
    average = student_grades.aggregate(value=Avg("value"))["value"]
    absence_count = Attendance.objects.filter(
        student=student,
        lesson__assignment=assignment,
    ).filter(
        Q(status=Attendance.Status.ABSENT)
        | Q(status=Attendance.Status.EXCUSED)
    ).count()

    return JsonResponse(
        {
            "saved": True,
            "average": round(average, 2) if average is not None else None,
            "absence_count": absence_count,
        }
    )


@login_required
def lesson_detail(request, lesson_id):
    lesson = get_object_or_404(
        Lesson.objects.select_related(
            "assignment__teacher",
            "assignment__group",
            "assignment__subject",
        ),
        pk=lesson_id,
    )
    can_edit = _can_edit_assignment(request.user, lesson.assignment)
    can_view = (
        can_edit
        or can_view_all_journals(request.user)
        or lesson.assignment.group.curator_id == request.user.id
    )
    if not can_view:
        raise PermissionDenied
    is_archived = _assignment_is_archived(lesson.assignment)

    students = list(
        Student.objects.filter(
            group=lesson.assignment.group,
            is_active=True,
        ).prefetch_related(
            Prefetch(
                "grades",
                queryset=Grade.objects.filter(lesson=lesson),
                to_attr="lesson_grades",
            ),
            Prefetch(
                "attendance_records",
                queryset=Attendance.objects.filter(lesson=lesson),
                to_attr="lesson_attendance",
            ),
        )
    )

    if request.method == "POST":
        if not can_edit:
            raise PermissionDenied
        archived_response = _reject_archived_change(
            request,
            lesson.assignment,
        )
        if archived_response:
            return archived_response
        reason = _change_reason_from_request(request)
        comment = request.POST.get("change_comment", "").strip()
        try:
            with transaction.atomic():
                for student in students:
                    grade_value = request.POST.get(
                        f"grade_{student.pk}",
                        "",
                    ).strip()
                    attendance_status = request.POST.get(
                        f"attendance_{student.pk}",
                        Attendance.Status.PRESENT,
                    )

                    _save_grade_result(
                        student,
                        lesson,
                        grade_value,
                        user=request.user,
                        reason=reason,
                        comment=comment,
                    )

                    if attendance_status in Attendance.Status.values:
                        _save_attendance(
                            student,
                            lesson,
                            attendance_status,
                            user=request.user,
                            reason=reason,
                            comment=comment,
                        )
        except ChangeReasonRequired as error:
            messages.error(request, str(error))
            return redirect("journal:lesson_detail", lesson_id=lesson.pk)

        messages.success(request, "Данные занятия сохранены.")
        return redirect("journal:lesson_detail", lesson_id=lesson.pk)

    rows = []
    for student in students:
        grade_record = student.lesson_grades[0] if student.lesson_grades else None
        grade_value = ""
        if grade_record:
            if grade_record.pass_result == Grade.PassResult.PASSED:
                grade_value = "З"
            elif grade_record.pass_result == Grade.PassResult.FAILED:
                grade_value = "НЗ"
            elif grade_record.value is not None:
                grade_value = grade_record.value
        rows.append(
            {
                "student": student,
                "grade": grade_value,
                "grade_exists": grade_record is not None,
                "attendance": student.lesson_attendance[0].status
                if student.lesson_attendance
                else Attendance.Status.PRESENT,
                "attendance_exists": bool(student.lesson_attendance),
            }
        )

    return render(
        request,
        "journal/lesson_detail.html",
        {
            "lesson": lesson,
            "rows": rows,
            "attendance_choices": Attendance.Status.choices,
            "grading_scheme": lesson.grading_scheme,
            "can_edit": can_edit,
            "is_archived": is_archived,
        },
    )


@login_required
def delete_lesson(request, lesson_id):
    lesson = get_object_or_404(
        Lesson.objects.select_related(
            "assignment__teacher",
            "assignment__group",
            "assignment__subject",
            "curriculum_item",
        ),
        pk=lesson_id,
    )
    if not _can_edit_assignment(request.user, lesson.assignment):
        raise PermissionDenied

    assignment_id = lesson.assignment_id
    if request.method == "POST":
        archived_response = _reject_archived_change(
            request,
            lesson.assignment,
        )
        if archived_response:
            return archived_response
        deleted_lesson_id = lesson.pk
        JournalChangeLog.objects.create(
            user=request.user,
            assignment=lesson.assignment,
            lesson_date=lesson.date,
            lesson_topic=lesson.topic,
            entity_type=JournalChangeLog.EntityType.LESSON,
            action=JournalChangeLog.Action.DELETE,
            old_value=str(lesson),
            new_value="",
            reason=request.POST.get(
                "change_reason",
                "Удаление занятия преподавателем",
            ).strip(),
            comment=request.POST.get("change_comment", "").strip(),
        )
        lesson.delete()
        logger.info(
            "lesson_deleted assignment_id=%s lesson_id=%s user_id=%s",
            assignment_id,
            deleted_lesson_id,
            request.user.pk,
        )
        messages.success(
            request,
            "Занятие удалено. Тема КТП снова доступна как запланированная.",
        )
        return redirect(
            "journal:assignment_detail",
            assignment_id=assignment_id,
        )

    return render(
        request,
        "journal/delete_lesson.html",
        {"lesson": lesson},
    )
