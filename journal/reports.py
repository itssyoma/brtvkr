import csv
from datetime import datetime
import hashlib
from io import BytesIO, StringIO
from pathlib import Path
import zipfile

from django.conf import settings
from django.utils import timezone

from .models import (
    Attendance,
    CollegeInformation,
    Grade,
    user_full_name,
)
from .permissions import user_role_name


def _safe_name(value):
    return "".join(
        character if character.isalnum() or character in "-_" else "_"
        for character in str(value)
    ).strip("_")


def _grade_text(grade):
    if grade is None:
        return ""
    if grade.pass_result == Grade.PassResult.PASSED:
        return "Зачёт"
    if grade.pass_result == Grade.PassResult.FAILED:
        return "Незачёт"
    return str(grade.value) if grade.value is not None else ""


def _student_summary_metrics(students, lessons, grades, attendance):
    metrics = {}
    for student in students:
        numeric_grades = []
        absence_count = 0
        for lesson in lessons:
            grade = grades.get((student.id, lesson.id))
            if grade and grade.value is not None:
                numeric_grades.append(grade.value)
            attendance_record = attendance.get((student.id, lesson.id))
            if attendance_record and attendance_record.status in {
                Attendance.Status.ABSENT,
                Attendance.Status.EXCUSED,
            }:
                absence_count += 1
        metrics[student.id] = {
            "average": (
                f"{sum(numeric_grades) / len(numeric_grades):.2f}".replace(
                    ".",
                    ",",
                )
                if numeric_grades
                else "—"
            ),
            "absence_count": absence_count,
        }
    return metrics


def _journal_rows(assignment):
    lessons = list(assignment.lessons.order_by("date", "id"))
    students = list(
        assignment.group.students.filter(is_active=True).order_by(
            "last_name",
            "first_name",
            "middle_name",
        )
    )
    grades = {
        (grade.student_id, grade.lesson_id): grade
        for grade in Grade.objects.filter(lesson__assignment=assignment)
    }
    attendance = {
        (record.student_id, record.lesson_id): record
        for record in Attendance.objects.filter(lesson__assignment=assignment)
    }
    for lesson in lessons:
        for student in students:
            attendance_record = attendance.get((student.id, lesson.id))
            yield {
                "date": lesson.date,
                "topic": lesson.topic,
                "lesson_type": lesson.get_lesson_type_display(),
                "student": str(student),
                "student_card": student.student_card_number,
                "grade": _grade_text(grades.get((student.id, lesson.id))),
                "attendance": (
                    attendance_record.get_status_display()
                    if attendance_record
                    else Attendance.Status.PRESENT.label
                ),
            }


def _build_csv(assignment):
    output = StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(
        [
            "Группа",
            "Дисциплина",
            "Учебный год",
            "Семестр",
            "Дата занятия",
            "Тема",
            "Вид занятия",
            "Студент",
            "Номер студенческого билета",
            "Оценка",
            "Посещаемость",
        ]
    )
    for row in _journal_rows(assignment):
        writer.writerow(
            [
                assignment.group.name,
                assignment.subject.name,
                assignment.academic_year,
                assignment.get_semester_display(),
                row["date"].isoformat(),
                row["topic"],
                row["lesson_type"],
                row["student"],
                row["student_card"],
                row["grade"],
                row["attendance"],
            ]
        )
    return output.getvalue().encode("utf-8")


def _register_pdf_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/Library/Fonts/Arial Unicode.ttf"),
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
    ]
    bold_candidates = [
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
    ]
    regular = next((path for path in candidates if path.exists()), None)
    bold = next((path for path in bold_candidates if path.exists()), regular)
    if regular is None:
        raise RuntimeError("Не найден шрифт с поддержкой кириллицы.")
    pdfmetrics.registerFont(TTFont("JournalSans", str(regular)))
    pdfmetrics.registerFont(TTFont("JournalSansBold", str(bold)))


def _build_pdf(assignment, generated_by):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        LongTable,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        TableStyle,
    )
    from xml.sax.saxutils import escape

    _register_pdf_fonts()
    college = CollegeInformation.get_current()
    lessons = list(
        assignment.lessons.select_related("curriculum_item").order_by(
            "date",
            "id",
        )
    )
    students = list(
        assignment.group.students.filter(is_active=True).order_by(
            "last_name",
            "first_name",
            "middle_name",
        )
    )
    grades = {
        (grade.student_id, grade.lesson_id): grade
        for grade in Grade.objects.filter(lesson__assignment=assignment)
    }
    attendance = {
        (record.student_id, record.lesson_id): record
        for record in Attendance.objects.filter(lesson__assignment=assignment)
    }
    attendance_marks = {
        Attendance.Status.PRESENT: "",
        Attendance.Status.ABSENT: "Н",
        Attendance.Status.EXCUSED: "У",
        Attendance.Status.LATE: "О",
    }
    student_metrics = _student_summary_metrics(
        students,
        lessons,
        grades,
        attendance,
    )

    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Электронный журнал — {assignment.subject}",
        author=str(generated_by),
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "JournalTitle",
        parent=styles["Title"],
        fontName="JournalSansBold",
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#17212b"),
    )
    body_style = ParagraphStyle(
        "JournalBody",
        parent=styles["BodyText"],
        fontName="JournalSans",
        fontSize=8,
        leading=10,
    )
    header_style = ParagraphStyle(
        "JournalHeader",
        parent=body_style,
        fontName="JournalSansBold",
        textColor=colors.white,
        alignment=1,
    )
    cell_center_style = ParagraphStyle(
        "JournalCellCenter",
        parent=body_style,
        alignment=1,
        fontSize=8,
        leading=9,
    )
    student_style = ParagraphStyle(
        "JournalStudent",
        parent=body_style,
        fontSize=7.5,
        leading=9,
    )
    section_style = ParagraphStyle(
        "JournalSection",
        parent=title_style,
        fontSize=12,
        leading=15,
        spaceAfter=3 * mm,
    )

    story = []
    if college:
        story.extend(
            [
                Paragraph(escape(college.name), section_style),
                Spacer(1, 2 * mm),
            ]
        )
    story.extend(
        [
            Paragraph("Электронный журнал успеваемости", title_style),
            Spacer(1, 3 * mm),
            Paragraph(
                (
                    f"<b>Дисциплина:</b> {assignment.subject}<br/>"
                    f"<b>Группа:</b> {assignment.group}<br/>"
                    f"<b>Преподаватель:</b> "
                    f"{user_full_name(assignment.teacher)}<br/>"
                    f"<b>Период:</b> {assignment.academic_year}, "
                    f"{assignment.get_semester_display()}<br/>"
                    f"<b>Сформирован:</b> "
                    f"{timezone.localtime():%d.%m.%Y %H:%M}<br/>"
                    f"<b>Администратор:</b> "
                    f"{user_full_name(generated_by)}"
                ),
                body_style,
            ),
            Spacer(1, 5 * mm),
        ]
    )

    story.append(Paragraph("Ведомость успеваемости и посещаемости", section_style))
    if not lessons:
        story.append(Paragraph("Проведённые занятия отсутствуют.", body_style))
    else:
        chunk_size = 10
        for chunk_start in range(0, len(lessons), chunk_size):
            lesson_chunk = lessons[chunk_start : chunk_start + chunk_size]
            if chunk_start:
                story.append(PageBreak())
            story.append(
                Paragraph(
                    (
                        f"Занятия {chunk_start + 1}–"
                        f"{chunk_start + len(lesson_chunk)} из {len(lessons)}"
                    ),
                    body_style,
                )
            )
            story.append(Spacer(1, 2 * mm))
            headers = [
                Paragraph("№", header_style),
                Paragraph("Студент", header_style),
            ]
            for offset, lesson in enumerate(lesson_chunk, start=chunk_start + 1):
                headers.append(
                    Paragraph(
                        f"{offset}<br/>{lesson.date:%d.%m}",
                        header_style,
                    )
                )
            headers.extend(
                [
                    Paragraph("Средний<br/>балл", header_style),
                    Paragraph("Пропуски", header_style),
                ]
            )
            table_data = [headers]
            cell_backgrounds = []
            for student_number, student in enumerate(students, start=1):
                row = [
                    Paragraph(str(student_number), cell_center_style),
                    Paragraph(escape(str(student)), student_style),
                ]
                for column_index, lesson in enumerate(lesson_chunk, start=2):
                    grade = grades.get((student.id, lesson.id))
                    attendance_record = attendance.get((student.id, lesson.id))
                    attendance_status = (
                        attendance_record.status
                        if attendance_record
                        else Attendance.Status.PRESENT
                    )
                    attendance_mark = attendance_marks[attendance_status]
                    grade_value = _grade_text(grade)
                    compact_grade = {
                        "Зачёт": "З",
                        "Незачёт": "НЗ",
                    }.get(grade_value, grade_value)
                    cell_text = " ".join(
                        value
                        for value in (compact_grade, attendance_mark)
                        if value
                    ) or "—"
                    row.append(Paragraph(cell_text, cell_center_style))
                    fill = {
                        Attendance.Status.ABSENT: colors.HexColor("#fff0f0"),
                        Attendance.Status.EXCUSED: colors.HexColor("#fff6dc"),
                        Attendance.Status.LATE: colors.HexColor("#f0edf9"),
                    }.get(attendance_status)
                    if fill:
                        cell_backgrounds.append(
                            (
                                "BACKGROUND",
                                (column_index, student_number),
                                (column_index, student_number),
                                fill,
                            )
                        )
                metrics = student_metrics[student.id]
                row.extend(
                    [
                        Paragraph(metrics["average"], cell_center_style),
                        Paragraph(
                            str(metrics["absence_count"]),
                            cell_center_style,
                        ),
                    ]
                )
                table_data.append(row)

            if not students:
                table_data.append(
                    [
                        Paragraph("—", cell_center_style),
                        Paragraph("В группе нет активных студентов", body_style),
                    ]
                    + [""] * (len(lesson_chunk) + 2)
                )
            table = LongTable(
                table_data,
                repeatRows=1,
                colWidths=(
                    [9 * mm, 54 * mm]
                    + [16 * mm] * len(lesson_chunk)
                    + [25 * mm, 20 * mm]
                ),
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#176b55")),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#b9c4cb")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                    + cell_backgrounds
                )
            )
            story.append(table)

    story.append(PageBreak())
    story.append(Paragraph("Перечень проведённых занятий", section_style))
    lesson_headers = [
        "№",
        "Дата",
        "Тема занятия",
        "Вид занятия",
        "Часы",
        "Литература",
    ]
    lesson_data = [
        [Paragraph(value, header_style) for value in lesson_headers]
    ]
    for number, lesson in enumerate(lessons, start=1):
        hours = (
            str(lesson.curriculum_item.hours).replace(".", ",")
            if lesson.curriculum_item
            else "—"
        )
        lesson_data.append(
            [
                Paragraph(str(number), cell_center_style),
                Paragraph(lesson.date.strftime("%d.%m.%Y"), cell_center_style),
                Paragraph(escape(lesson.topic), body_style),
                Paragraph(escape(lesson.get_lesson_type_display()), body_style),
                Paragraph(hours, cell_center_style),
                Paragraph(
                    escape(lesson.curriculum_item.literature)
                    if lesson.curriculum_item
                    and lesson.curriculum_item.literature
                    else "—",
                    body_style,
                ),
            ]
        )
    if len(lesson_data) == 1:
        lesson_data.append(
            [
                Paragraph("—", cell_center_style),
                Paragraph("—", cell_center_style),
                Paragraph("Проведённые занятия отсутствуют", body_style),
                "",
                "",
                "",
            ]
        )
    lesson_table = LongTable(
        lesson_data,
        repeatRows=1,
        colWidths=[10 * mm, 25 * mm, 105 * mm, 48 * mm, 18 * mm, 48 * mm],
    )
    lesson_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#176b55")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#b9c4cb")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(lesson_table)

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("JournalSans", 8)
        canvas.setFillColor(colors.HexColor("#5c6975"))
        canvas.drawRightString(
            landscape(A4)[0] - 10 * mm,
            6 * mm,
            f"Страница {doc.page}",
        )
        canvas.restoreState()

    document.build(
        story,
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )
    return output.getvalue()


def _build_audit_xlsx(assignment):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Журнал изменений"
    headers = [
        "Дата и время",
        "Пользователь",
        "Роль",
        "Группа",
        "Дисциплина",
        "Дата занятия",
        "Тема занятия",
        "Студент",
        "Действие",
        "Объект",
        "Было",
        "Стало",
        "Основание",
        "Комментарий",
    ]
    worksheet.append(headers)
    for log in assignment.change_logs.select_related("user", "student"):
        role = user_role_name(log.user)
        worksheet.append(
            [
                timezone.localtime(log.created_at).replace(tzinfo=None),
                user_full_name(log.user),
                role,
                assignment.group.name,
                assignment.subject.name,
                log.lesson_date,
                log.lesson_topic,
                log.student_name,
                log.get_action_display(),
                log.get_entity_type_display(),
                log.old_value,
                log.new_value,
                log.reason,
                log.comment,
            ]
        )
    header_fill = PatternFill("solid", fgColor="176B55")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [19, 24, 16, 14, 28, 15, 42, 28, 16, 18, 18, 18, 42, 48]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_report_package(assignment, generated_by):
    base_name = "_".join(
        [
            "journal",
            _safe_name(assignment.group.name),
            _safe_name(assignment.subject.short_name or assignment.subject.name),
            datetime.now().strftime("%Y%m%d_%H%M"),
        ]
    )
    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(f"{base_name}.pdf", _build_pdf(assignment, generated_by))
        archive.writestr(f"{base_name}.csv", _build_csv(assignment))
        archive.writestr(
            f"{base_name}_changes.xlsx",
            _build_audit_xlsx(assignment),
        )
    return output.getvalue(), f"{base_name}.zip"


def build_academic_year_archive(academic_year, generated_by):
    from .models import TeachingAssignment

    assignments = list(
        TeachingAssignment.objects.filter(
            academic_year=academic_year,
        ).select_related("group", "subject", "teacher")
    )
    if not assignments:
        raise ValueError(
            "Для выбранного учебного года нет назначений дисциплин."
        )

    manifest = StringIO()
    writer = csv.writer(manifest, delimiter=";", lineterminator="\n")
    manifest.write("\ufeff")
    writer.writerow(
        [
            "Учебный год",
            "Группа",
            "Дисциплина",
            "Семестр",
            "Преподаватель",
            "Количество занятий",
        ]
    )

    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for assignment in assignments:
            package, package_name = build_report_package(
                assignment,
                generated_by,
            )
            folder = _safe_name(assignment.group.name) or "group"
            archive.writestr(
                f"{folder}/semester_{assignment.semester}/{package_name}",
                package,
            )
            writer.writerow(
                [
                    academic_year,
                    assignment.group.name,
                    assignment.subject.name,
                    assignment.get_semester_display(),
                    user_full_name(assignment.teacher),
                    assignment.lessons.count(),
                ]
            )
        archive.writestr(
            "registry.csv",
            manifest.getvalue().encode("utf-8"),
        )

    content = output.getvalue()
    filename = (
        f"academic_year_{_safe_name(academic_year)}_"
        f"{datetime.now():%Y%m%d_%H%M}.zip"
    )
    return content, filename, hashlib.sha256(content).hexdigest()
